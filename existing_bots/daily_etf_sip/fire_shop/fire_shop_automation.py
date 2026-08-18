#!/usr/bin/env python3
"""
FIRE Shop 3.0 - Daily ETF Buy Automation
Fetches live CMP & 20 DMA from NSE, ranks ETFs by dip, logs buy into Excel, sends email.

Usage:
    python fire_shop_automation.py --xlsx <path_to_xlsx> --email <your@email.com>
    python fire_shop_automation.py --xlsx FIRE_shop.xlsx --email you@gmail.com --smtp-host smtp.gmail.com --smtp-port 587 --smtp-user you@gmail.com --smtp-pass <app_password>
"""

import argparse
import json
import os
import smtplib
import sys
import time
from datetime import datetime, date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — edit defaults here or pass via CLI args
# ──────────────────────────────────────────────────────────────────────────────
DEFAULT_XLSX = "FIRE_shop_3_0_with_Jewels_-_BID_Investment_-_24_JUL_2024_Balkrishna.xlsx"

MASTER_ETFS = [
    ("NSE:CPSEETF", "CPSE ETF"), ("NSE:TOP100CASE", "NIFTY 100"),
    ("NSE:ESG", "NIFTY 100 ESG SECTOR LEADERS"), ("NSE:LOWVOLIETF", "NIFTY 100 Low Volatility 30"),
    ("NSE:HDFCQUAL", "NIFTY 100 Quality 30"), ("NSE:ALPHAETF", "NIFTY 200 Alpha 30"),
    ("NSE:MOM30IETF", "NIFTY 200 Momentum 30"), ("NSE:QUAL30IETF", "NIFTY 200 Quality 30"),
    ("NSE:NIFTYBEES", "NIFTY 50"), ("NSE:EQUAL50ADD", "NIFTY 50 Equal Weight"),
    ("NSE:NV20IETF", "NIFTY 50 Value 20"), ("NSE:MONIFTY500", "NIFTY 500"),
    ("NSE:ALPHA", "NIFTY Alpha 50"), ("NSE:ALPL30IETF", "NIFTY Alpha Low-Volatility 30"),
    ("NSE:AUTOIETF", "NIFTY Auto"), ("NSE:BANKBEES", "NIFTY Bank"),
    ("NSE:COMMOIETF", "NIFTY Commodities"), ("NSE:DIVOPPBEES", "NIFTY Dividend Opportunities 50 TRI"),
    ("NSE:EVINDIA", "Nifty EV and New Age Automotive"), ("NSE:BFSI", "NIFTY Financial Services"),
    ("NSE:FINIETF", "NIFTY Financial Services Ex-Bank"), ("NSE:FMCGIETF", "NIFTY FMCG"),
    ("NSE:HEALTHY", "NIFTY Healthcare"), ("NSE:CONSUMBEES", "NIFTY India Consumption"),
    ("NSE:TNIDETF", "NIFTY India Digital"), ("NSE:MAKEINDIA", "NIFTY India Manufacturing Total Return"),
    ("NSE:INFRAIETF", "NIFTY Infrastructure"),
    ("NSE:LICNMID100", "NIFTY Midcap 100"), ("NSE:MIDCAPETF", "NIFTY Midcap 150"),
    ("NSE:MIDCAP", "NIFTY Midcap 50"), ("NSE:MIDSMALL", "NIFTY MidSmallcap 400 Momentum Quality 100"),
    ("NSE:MNC", "NIFTY MNC"), ("NSE:NEXT50IETF", "NIFTY NEXT 50"),
    ("NSE:PHARMABEES", "NIFTY Pharma"), ("NSE:PVTBANIETF", "NIFTY Private Bank"),
    ("NSE:PSUBNKBEES", "NIFTY PSU Bank"), ("NSE:MOREALTY", "Nifty Realty"),
    ("NSE:MOSMALL250", "NIFTY Smallcap 250"), ("NSE:SMALLCAP", "NIFTY Smallcap 250 Momentum Quality 100"),
    ("NSE:BSE500IETF", "S&P BSE 500"), ("NSE:ICICIB22", "S&P BSE BHARAT 22"),
    ("NSE:MOVALUE", "S&P BSE Enhanced Value"), ("NSE:MOHEALTH", "S&P BSE Healthcare"),
    ("NSE:MIDSELIETF", "S&P BSE Midcap Select"), ("NSE:HDFCSENSEX", "Sensex"),
]

JEWELLERY_ETFS = [
    ("NSE:GOLDBEES", "Nippon India ETF Gold BeES"),
    ("NSE:SILVERBEES", "Nippon India Silver ETF"),
]

TOP_STOCKS = [
    ("NSE:RELIANCE", "Reliance Industries Ltd"), ("NSE:TCS", "Tata Consultancy Services Ltd"),
    ("NSE:HDFCBANK", "HDFC Bank Ltd"), ("NSE:ITC", "ITC Ltd"),
    ("NSE:SBIN", "State Bank of India"), ("NSE:BHARTIARTL", "Bharti Airtel Ltd"),
    ("NSE:BAJFINANCE", "Bajaj Finance Ltd"), ("NSE:LICI", "Life Insurance Corporation of India"),
    ("NSE:LT", "Larsen and Toubro Ltd"), ("NSE:ASIANPAINT", "Asian Paints Ltd"),
]

# ──────────────────────────────────────────────────────────────────────────────
# YAHOO FINANCE DATA FETCHING
# ──────────────────────────────────────────────────────────────────────────────

def get_nse_session():
    """Returns a dummy session object (kept for API compatibility)."""
    return requests.Session()


def nse_to_yahoo(symbol):
    """Convert NSE:AUTOIETF → AUTOIETF.NS for Yahoo Finance."""
    return symbol.replace("NSE:", "") + ".NS"


def compute_rsi(closes, period=14):
    """
    RSI matching fire_shop backtest: simple rolling mean of gains/losses.
    Needs period+1 closes; returns None if insufficient history.
    """
    if closes is None or len(closes) < period + 1:
        return None
    gains = []
    losses = []
    for i in range(1, len(closes)):
        delta = closes[i] - closes[i - 1]
        gains.append(max(delta, 0.0))
        losses.append(max(-delta, 0.0))
    avg_gain = sum(gains[-period:]) / period
    avg_loss = sum(losses[-period:]) / period
    if avg_loss == 0:
        return 100.0 if avg_gain > 0 else 50.0
    rs = avg_gain / avg_loss
    return round(100.0 - (100.0 / (1.0 + rs)), 2)


def fetch_etf_data(session, symbol, rsi_period=14):
    """
    Fetch CMP, 20 DMA, avg daily volume, and RSI via Yahoo Finance.
    Returns (cmp, dma20, avg_volume, rsi) or (None, None, None, None).
    """
    ticker = nse_to_yahoo(symbol)
    # 3mo gives enough bars for stable RSI(14) + 20 DMA
    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}"
           f"?interval=1d&range=3mo")
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        r = requests.get(url, headers=headers, timeout=10)
        data = r.json()
        result = data["chart"]["result"][0]
        quotes  = result["indicators"]["quote"][0]
        closes  = [c for c in quotes["close"]  if c is not None]
        volumes = [v for v in quotes.get("volume", []) if v is not None]
        if not closes:
            return None, None, None, None
        cmp        = closes[-1]
        dma20      = sum(closes[-20:]) / min(20, len(closes[-20:]))
        avg_volume = sum(volumes[-20:]) / min(20, len(volumes[-20:])) if volumes else 0
        rsi        = compute_rsi(closes, period=rsi_period)
        return round(cmp, 2), round(dma20, 2), int(avg_volume), rsi
    except Exception as e:
        print(f"      [fetch error] {symbol}: {e}")
        return None, None, None, None


def compute_pct_change(cmp, dma20):
    if cmp and dma20 and dma20 != 0:
        return (cmp - dma20) / dma20
    return None


# ──────────────────────────────────────────────────────────────────────────────
# RANKING
# ──────────────────────────────────────────────────────────────────────────────

MIN_VOLUME = 20000      # exclude ETFs with avg daily volume below this
DEFAULT_BUY_RANK_MODE = "rsi"  # lowest RSI(14); "dma" = deepest 20DMA dip
DEFAULT_RSI_PERIOD = 14

# ── Accumulation guard ────────────────────────────────────────────────────────
MAX_INVESTED_PER_ETF  = 15000   # ₹ — once invested >= this, apply cooldown
AVG_DOWN_COOLDOWN_DAYS = 7      # days — min gap between buys after threshold hit

def rank_instruments(instruments, session, label, rank_mode=None, rsi_period=None):
    """
    Fetch live data, then:
    1. Filter by avg daily volume (liquidity)
    2. Rank by lowest RSI(14) (default) or deepest dip vs 20 DMA
    """
    mode = (rank_mode or DEFAULT_BUY_RANK_MODE).lower()
    period = int(rsi_period or DEFAULT_RSI_PERIOD)
    if mode not in ("rsi", "dma"):
        raise ValueError(f"unknown rank_mode: {mode}")

    print(f"\n  Fetching {label} ({len(instruments)} instruments) "
          f"[rank={mode.upper()}" + (f"/{period}" if mode == "rsi" else "") + "]...")
    results = []
    for code, name in instruments:
        cmp, dma20, avg_vol, rsi = fetch_etf_data(session, code, rsi_period=period)
        pct = compute_pct_change(cmp, dma20)
        if cmp is None or (mode == "rsi" and rsi is None) or (mode == "dma" and pct is None):
            status = "fetch failed"
        else:
            status = (
                f"CMP=₹{cmp}, RSI={rsi}, 20DMA=₹{dma20}, Δ={pct:.2%}, Vol={avg_vol:,}"
                if pct is not None
                else f"CMP=₹{cmp}, RSI={rsi}, Vol={avg_vol:,}"
            )
            results.append({
                "code": code,
                "name": name,
                "cmp": cmp,
                "dma20": dma20,
                "pct": pct if pct is not None else 0.0,
                "rsi": rsi,
                "avg_volume": avg_vol,
            })
        print(f"    {code:<25} {status}")
        time.sleep(0.2)

    # Step 1 — exclude ETFs below minimum volume threshold
    top_by_volume = [r for r in results if r["avg_volume"] >= MIN_VOLUME]
    excluded      = [r["code"] for r in results if r["avg_volume"] < MIN_VOLUME]
    if excluded:
        print(f"\n  ⛔ Low volume excluded: {', '.join(excluded)}")
    print(
        f"  ✅ {len(top_by_volume)} ETFs above volume threshold "
        f"shortlisted for {mode.upper()} ranking"
    )

    # Step 2 — rank: lowest RSI or deepest (most negative) DMA dip
    if mode == "rsi":
        top_by_volume.sort(key=lambda x: x["rsi"])
    else:
        top_by_volume.sort(key=lambda x: x["pct"])
    for i, r in enumerate(top_by_volume):
        r["rank"] = i + 1
    return top_by_volume


# ──────────────────────────────────────────────────────────────────────────────
# HOLDINGS FILTER — read Current Holdings and decide what to suppress
# ──────────────────────────────────────────────────────────────────────────────

def load_current_holdings(xlsx_path):
    """
    Reads Current Holdings sheet and returns a dict:
      { "NSE:AUTOIETF": { "avg_price": 24.95, "next_bid": 24.20, "shop": "etf" }, ... }

    A holding is SUPPRESSED from the shopping list unless:
      (a) it has been sold  → not in this dict at all
      (b) CMP has dropped to next BID level (avg_price * (1 + bid_pct)) → show again for averaging down
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Current Holdings"]

    # Column map (1-indexed): N=14 ShopType, O=15 BuyDate, P=16 NSECode,
    # Q=17 Name, R=18 CMP, E=5 BuyPrice(AvgPrice col J=10), Q=17 NextBIDlevel col Q=17
    # From the sheet: col E=BuyPrice, col J=AvgPrice, col Q=NextBIDlevel (17th col)
    COL_SHOP     = 1   # A  Shop Type
    COL_CODE     = 3   # C  NSE Code
    COL_AVG      = 10  # J  Avg Price
    COL_NEXT_BID = 18  # R  Next BID price

    holdings = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        code = row[COL_CODE - 1]
        if not code or not isinstance(code, str) or not code.startswith("NSE:"):
            continue
        shop_raw = row[COL_SHOP - 1] or ""
        if "ETF" in shop_raw.upper() and "JEWEL" in shop_raw.upper():
            shop = "jewellery"
        elif "STOCK" in shop_raw.upper():
            shop = "stock"
        else:
            shop = "etf"

        avg_price  = row[COL_AVG - 1]
        next_bid   = row[COL_NEXT_BID - 1]
        target     = row[11 - 1]    # col K = Target Price
        name       = row[4 - 1]     # col D = Underlying Asset name
        buy_price  = row[5 - 1]     # col E = Buy Price
        buy_qty    = row[6 - 1]     # col F = Actual Buy Qty

        # Parse buy date
        raw_date = row[2 - 1]   # col B = Buy Date
        try:
            from datetime import datetime as _dt
            if isinstance(raw_date, _dt):
                buy_date = raw_date.date()
            elif isinstance(raw_date, str):
                for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        buy_date = _dt.strptime(raw_date, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    buy_date = None
            else:
                buy_date = None
        except Exception:
            buy_date = None

        # Accumulate total qty and total invested across multiple buy rows
        if code not in holdings:
            holdings[code] = {
                "avg_price":      avg_price,
                "next_bid":       next_bid,
                "target":         target,
                "shop":           shop,
                "name":           name,
                "total_qty":      float(buy_qty or 0),
                "total_invested": round(float(buy_price or 0) * float(buy_qty or 0), 2),
                "last_buy_date":  buy_date,
            }
        else:
            # Update avg and next_bid from latest row (has cumulative avg)
            if avg_price:
                holdings[code]["avg_price"] = avg_price
            if next_bid and (not holdings[code]["next_bid"] or next_bid < holdings[code]["next_bid"]):
                holdings[code]["next_bid"] = next_bid
            # Accumulate qty and invested
            holdings[code]["total_qty"]      += float(buy_qty or 0)
            holdings[code]["total_invested"] += round(float(buy_price or 0) * float(buy_qty or 0), 2)
            # Track latest buy date
            if buy_date and (not holdings[code]["last_buy_date"] or buy_date > holdings[code]["last_buy_date"]):
                holdings[code]["last_buy_date"] = buy_date

    return holdings


def load_sold_codes(xlsx_path):
    """Return set of NSE codes that appear in the Sold sheet (fully exited)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Sold"]
    sold = set()
    # Col C = NSE Code in Sold sheet (row 3 onward)
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = row[2]   # column C
        if code and isinstance(code, str) and code.startswith("NSE:"):
            sold.add(code)
    return sold


def apply_holdings_filter(ranked_list, holdings, sold_codes, live_cmp_map):
    """
    Filter a ranked list based on holding status.

    Rules:
      1. Not in holdings → show normally ✅
      2. In holdings AND still held (not fully sold) AND CMP > next_bid → HIDE ❌
      3. In holdings AND CMP <= next_bid → show with tag "🔁 AVG DOWN" ✅
      4. Was in holdings but now in sold → show normally ✅ (fully exited)

    Returns filtered list with optional 'note' field added.
    """
    filtered = []
    for item in ranked_list:
        code = item["code"]
        cmp  = item["cmp"]

        if code not in holdings:
            # Not in current holdings — never bought or fully sold, show normally
            filtered.append({**item, "note": ""})
            continue

        # Currently held
        h        = holdings[code]
        next_bid = h.get("next_bid")

        if next_bid and cmp is not None and cmp <= next_bid:
            # Price has dipped to next BID — potential AVG DOWN
            # But check accumulation guard first
            total_invested = h.get("total_invested", 0)
            last_buy_date  = h.get("last_buy_date")

            if total_invested >= MAX_INVESTED_PER_ETF and last_buy_date:
                from datetime import date as _date
                days_since = (_date.today() - last_buy_date).days
                if days_since < AVG_DOWN_COOLDOWN_DAYS:
                    print(f"    🛑 {code} cooldown: ₹{total_invested:,.0f} invested, "
                          f"last buy {days_since}d ago (need {AVG_DOWN_COOLDOWN_DAYS}d gap)")
                    continue   # skip — too soon to buy again
                else:
                    note = (f"🔁 AVG DOWN (next BID: ₹{next_bid}) | "
                            f"₹{total_invested:,.0f} invested | last buy {days_since}d ago")
            else:
                note = f"🔁 AVG DOWN (next BID: ₹{next_bid})"

            filtered.append({**item, "note": note})
        else:
            # Still holding, price hasn't dipped enough — suppress
            print(f"    ⛔ {code} suppressed (already holding, CMP ₹{cmp} > next BID ₹{next_bid})")

    return filtered


# ──────────────────────────────────────────────────────────────────────────────
# EXCEL LOGGING
# ──────────────────────────────────────────────────────────────────────────────

BUY_COLOR = "E2EFDA"   # light green
HEADER_COLOR = "4472C4"
TODAY_COLOR = "FFF2CC"

def log_to_excel(xlsx_path, today_picks, all_ranked, investment_per_tx):
    """
    Appends today's recommended buys to the Current Holdings sheet
    and updates the Shopping list sheet with live rankings.
    """
    wb = openpyxl.load_workbook(xlsx_path)

    # ── 1. Append to Current Holdings ────────────────────────────────────────
    ws_hold = wb["Current Holdings"]
    today_str = datetime.now().strftime("%d/%m/%Y")

    # Find first empty row after row 6 (data starts at row 7 in original)
    last_row = ws_hold.max_row + 1
    for row_idx in range(7, ws_hold.max_row + 2):
        if ws_hold.cell(row=row_idx, column=2).value is None:
            last_row = row_idx
            break

    shop_type_map = {"etf": "Equity ETF Buy", "jewellery": "Jewellery ETF Buy", "stock": "Stock Buy"}

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", start_color=TODAY_COLOR)

    for pick in today_picks:
        cmp       = pick["cmp"] if pick["cmp"] else 0
        qty       = int(investment_per_tx // cmp) if cmp else 1
        invested  = round(qty * cmp, 2)
        shop_label = shop_type_map.get(pick["shop"], "Equity ETF Buy")

        # Calculate avg price considering existing holdings of same code
        existing_qty      = 0
        existing_invested = 0
        for row in ws_hold.iter_rows(min_row=7, values_only=True):
            if row[2] == pick["code"] and row[4] and row[5]:
                existing_qty      += float(row[5] or 0)
                existing_invested += float(row[6] or 0)

        total_qty      = existing_qty + qty
        total_invested = existing_invested + invested
        avg_price      = round(total_invested / total_qty, 4) if total_qty else cmp

        # BID level from strategy: -3% for ETFs
        bid_level      = -0.03
        next_bid_price = round(avg_price * (1 + bid_level), 4)

        row_data = [
            shop_label,      # Col 1:  Shop Type
            today_str,       # Col 2:  Buy Date
            pick["code"],    # Col 3:  NSE Code
            pick["name"],    # Col 4:  Underlying Asset
            cmp,             # Col 5:  Buy Price
            qty,             # Col 6:  Actual Buy Qty
            invested,        # Col 7:  Invested amount
            total_qty,       # Col 8:  Total Qty
            total_invested,  # Col 9:  Total Invested
            avg_price,       # Col 10: Avg Price
            None,            # Col 11: Target Price (leave blank)
            cmp,             # Col 12: CMP
            None,            # Col 13: Sell Date
            None,            # Col 14: Total Invested amt
            None,            # Col 15: Overall Notional P/L
            None,            # Col 16: Notional P/L%
            bid_level,       # Col 17: Next BID level
            next_bid_price,  # Col 18: Next BID price
        ]
        for col_idx, val in enumerate(row_data, start=1):
            if val is not None:
                cell = ws_hold.cell(row=last_row, column=col_idx, value=val)
                cell.border = border
                cell.fill = fill
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal="center")

        last_row += 1

    # ── 2. Update Shopping list with live rankings ────────────────────────────
    ws_shop = wb["Shopping list"]

    def write_ranked_block(ws, start_row, ranked, top_n=5):
        """Overwrite a ranked block in shopping list."""
        for i, item in enumerate(ranked[:top_n]):
            r = start_row + i
            ws.cell(row=r, column=1, value=i + 1)
            ws.cell(row=r, column=2, value=round(item["pct"], 6))
            ws.cell(row=r, column=3, value=item["code"])
            ws.cell(row=r, column=4, value=item["name"])
            ws.cell(row=r, column=5, value=item["cmp"])
            for col in range(1, 6):
                ws.cell(row=r, column=col).font = Font(name="Arial", size=10)

    # ETFs start at row 6 in filtered block; Jewellery at row 15; Stocks at row 21
    if all_ranked.get("etf"):
        write_ranked_block(ws_shop, 6, all_ranked["etf"], top_n=5)
    if all_ranked.get("jewellery"):
        write_ranked_block(ws_shop, 15, all_ranked["jewellery"], top_n=2)
    if all_ranked.get("stock"):
        write_ranked_block(ws_shop, 21, all_ranked["stock"], top_n=3)

    wb.save(xlsx_path)   # overwrite original file
    print(f"\n  ✅ Excel saved: {xlsx_path}")
    return xlsx_path


# ──────────────────────────────────────────────────────────────────────────────
# EMAIL
# ──────────────────────────────────────────────────────────────────────────────

def build_email_html(today_picks, all_ranked, investment_per_tx):
    date_str = datetime.now().strftime("%d %b %Y")

    def pick_rows(picks):
        if not picks:
            return '<tr><td colspan="8" style="padding:12px;text-align:center;color:#888">⚠️ No new picks today — all top instruments are currently held. They reappear when CMP hits the Next BID level or after selling.</td></tr>'
        rows = ""
        for p in picks:
            qty = int(investment_per_tx // p["cmp"]) if p["cmp"] else "—"
            invested = round(qty * p["cmp"], 2) if isinstance(qty, int) else "—"
            note = p.get("note", "")
            note_color = "#e67e00" if "AVG DOWN" in note else "#27ae60"
            rows += f"""
            <tr>
              <td style="padding:8px;border:1px solid #ddd">{p.get('rank','')}</td>
              <td style="padding:8px;border:1px solid #ddd"><b>{p['code']}</b></td>
              <td style="padding:8px;border:1px solid #ddd">{p['name']}</td>
              <td style="padding:8px;border:1px solid #ddd">₹{p['cmp']}</td>
              <td style="padding:8px;border:1px solid #ddd">{p['pct']:.2%}</td>
              <td style="padding:8px;border:1px solid #ddd">{qty}</td>
              <td style="padding:8px;border:1px solid #ddd">₹{invested}</td>
              <td style="padding:8px;border:1px solid #ddd;color:{note_color}">{note or "🆕 New Buy"}</td>
            </tr>"""
        return rows

    def ranked_table(ranked, label, top_n=5):
        rows = ""
        for item in ranked[:top_n]:
            rows += f"""
            <tr>
              <td style="padding:6px;border:1px solid #eee">{item['rank']}</td>
              <td style="padding:6px;border:1px solid #eee">{item['code']}</td>
              <td style="padding:6px;border:1px solid #eee">{item['name']}</td>
              <td style="padding:6px;border:1px solid #eee">₹{item['cmp']}</td>
              <td style="padding:6px;border:1px solid #eee">{item['pct']:.2%}</td>
            </tr>"""
        return f"""
        <h3 style="margin-top:24px;color:#333">{label}</h3>
        <table style="border-collapse:collapse;width:100%;font-size:13px">
          <tr style="background:#4472C4;color:white">
            <th style="padding:8px">#</th><th style="padding:8px">Code</th>
            <th style="padding:8px">Name</th><th style="padding:8px">CMP</th>
            <th style="padding:8px">Δ 20DMA</th>
          </tr>{rows}
        </table>"""

    today_section = f"""
    <h2 style="color:#2E75B6">📋 Today's Recommended Buys — {date_str}</h2>
    <p>Capital per transaction: <b>₹{investment_per_tx:,.0f}</b></p>
    <table style="border-collapse:collapse;width:100%;font-size:14px">
      <tr style="background:#2E75B6;color:white">
        <th style="padding:10px">#</th><th style="padding:10px">Code</th>
        <th style="padding:10px">Name</th><th style="padding:10px">CMP</th>
        <th style="padding:10px">Δ 20DMA</th><th style="padding:10px">Qty</th>
        <th style="padding:10px">Invest</th>
        <th style="padding:10px">Type</th>
      </tr>
      {pick_rows(today_picks)}
    </table>"""

    full_rankings = "".join([
        ranked_table(all_ranked.get("etf", []), "🏦 Equity ETFs — Full Ranking", 10),
        ranked_table(all_ranked.get("jewellery", []), "💍 Jewellery ETFs", 2),
        ranked_table(all_ranked.get("stock", []), "📈 Top Stocks", 5),
    ])

    return f"""
    <html><body style="font-family:Arial,sans-serif;max-width:900px;margin:auto;padding:20px">
      <h1 style="color:#1F497D">🔥 FIRE Shop 3.0 — Daily Buy Signal</h1>
      <p style="color:#888">Generated on {datetime.now().strftime('%d %b %Y %H:%M IST')}</p>
      <hr/>
      {today_section}
      <hr/>
      <h2 style="color:#2E75B6">📊 Full Rankings (Live)</h2>
      {full_rankings}
      <hr/>
      <p style="color:#aaa;font-size:11px">
        This is an automated signal based on your FIRE Shop 3.0 strategy.<br/>
        Always verify CMP on your broker before placing the order.
      </p>
    </body></html>"""


def send_email(to_email, html_body, smtp_host, smtp_port, smtp_user, smtp_pass, xlsx_path=None):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"🔥 FIRE Shop — Today's Buy Signal {datetime.now().strftime('%d %b %Y')}"
    msg["From"] = smtp_user
    msg["To"] = to_email
    msg.attach(MIMEText(html_body, "html"))

    if xlsx_path and os.path.exists(xlsx_path):
        from email.mime.base import MIMEBase
        from email import encoders
        with open(xlsx_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(xlsx_path)}"')
        msg.attach(part)

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.ehlo()
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(smtp_user, to_email, msg.as_string())
    print(f"  ✅ Email sent to {to_email}")


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def load_investment_per_tx(xlsx_path):
    """Read investment per transaction from Current Holdings sheet."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["Current Holdings"]
        val = ws.cell(row=2, column=2).value   # "Investment per transaction (40 parts)"
        return float(val) if val else 3000
    except Exception:
        return 3000


def main():
    parser = argparse.ArgumentParser(description="FIRE Shop 3.0 Daily Buy Automation")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Path to your FIRE shop Excel file")
    parser.add_argument("--email", required=False, help="Email address to send the signal to")
    parser.add_argument("--smtp-host", default="smtp.gmail.com")
    parser.add_argument("--smtp-port", type=int, default=587)
    parser.add_argument("--smtp-user", default=None, help="SMTP login (your Gmail)")
    parser.add_argument("--smtp-pass", default=None, help="SMTP password / App password")
    parser.add_argument("--dry-run", action="store_true", help="Fetch data but don't write Excel or send email")
    parser.add_argument("--top-etfs", type=int, default=1, help="How many ETFs to pick from ETF shop (default 1)")
    args = parser.parse_args()

    print("=" * 60)
    print(f"  FIRE Shop 3.0 — Daily Run  [{datetime.now().strftime('%d %b %Y %H:%M')}]")
    print("=" * 60)

    investment_per_tx = load_investment_per_tx(args.xlsx)
    print(f"\n  Investment per transaction: ₹{investment_per_tx:,.0f}")

    session = get_nse_session()

    # Fetch & rank all shops
    etf_ranked    = rank_instruments(MASTER_ETFS,    session, "Equity ETFs")
    jewel_ranked  = rank_instruments(JEWELLERY_ETFS, session, "Jewellery ETFs")
    stock_ranked  = rank_instruments(TOP_STOCKS,     session, "Stocks")

    all_ranked_raw = {"etf": etf_ranked, "jewellery": jewel_ranked, "stock": stock_ranked}

    # ── Holdings filter ───────────────────────────────────────────────────────
    print("\n  Loading current holdings & sold positions...")
    holdings   = load_current_holdings(args.xlsx)
    sold_codes = load_sold_codes(args.xlsx)

    print(f"  Currently holding {len(holdings)} positions: {', '.join(holdings.keys()) or 'none'}")


    live_cmp_map = {r["code"]: r["cmp"] for group in all_ranked_raw.values() for r in group}

    print("\n  Applying holdings filter...")
    etf_filtered   = apply_holdings_filter(etf_ranked,   holdings, sold_codes, live_cmp_map)
    jewel_filtered = apply_holdings_filter(jewel_ranked, holdings, sold_codes, live_cmp_map)
    stock_filtered = apply_holdings_filter(stock_ranked, holdings, sold_codes, live_cmp_map)

    all_ranked = {"etf": etf_filtered, "jewellery": jewel_filtered, "stock": stock_filtered}

    # Pick today's buys: top 1 from each shop (most dipped, after filter)
    today_picks = []
    for shop_key, ranked, n in [("etf", etf_filtered, args.top_etfs),
                                 ("jewellery", jewel_filtered, 1),
                                 ("stock", stock_filtered, 1)]:
        for item in ranked[:n]:
            today_picks.append({**item, "shop": shop_key})

    print("\n" + "=" * 60)
    print("  TODAY'S BUY PICKS")
    print("=" * 60)
    if not today_picks:
        print("  ⚠️  No picks today — all top-ranked instruments are already held.")
        print("     They will reappear when CMP drops to their Next BID level,")
        print("     or once they are sold.")
    for p in today_picks:
        qty = int(investment_per_tx // p["cmp"]) if p["cmp"] else "?"
        note = f"  ← {p['note']}" if p.get("note") else ""
        print(f"  [{p['shop'].upper():10}] {p['code']:<25} CMP=₹{p['cmp']:<8} "
              f"Δ20DMA={p['pct']:+.2%}  Qty={qty}  "
              f"Invest=₹{round(qty*p['cmp'],2) if isinstance(qty,int) else '?'}{note}")

    if not args.dry_run:
        # Write to Excel
        updated_xlsx = log_to_excel(args.xlsx, today_picks, all_ranked, investment_per_tx)

        # Send email
        if args.email and args.smtp_user and args.smtp_pass:
            html = build_email_html(today_picks, all_ranked, investment_per_tx)
            send_email(
                to_email=args.email,
                html_body=html,
                smtp_host=args.smtp_host,
                smtp_port=args.smtp_port,
                smtp_user=args.smtp_user,
                smtp_pass=args.smtp_pass,
                xlsx_path=updated_xlsx,
            )
        elif args.email:
            print("\n  [WARN] Email address provided but SMTP credentials missing. Skipping email.")
            html = build_email_html(today_picks, all_ranked, investment_per_tx)
            preview_path = f"/tmp/fire_signal_{date.today().isoformat()}.html"
            with open(preview_path, "w") as f:
                f.write(html)
            print(f"  📄 Email preview saved to: {preview_path}")
    else:
        print("\n  [DRY RUN] No files written, no email sent.")

    print("\n✅ Done.")


if __name__ == "__main__":
    main()
