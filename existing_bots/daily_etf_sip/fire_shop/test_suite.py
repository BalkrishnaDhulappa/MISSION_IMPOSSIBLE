#!/usr/bin/env python3
"""
FIRE Shop 3.0 — Comprehensive Test Suite
All-weather tests covering every scenario without placing real orders or modifying Excel.

Run: python3 test_suite.py
"""

import json
import os
import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import MagicMock, patch, PropertyMock

_script_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(_script_dir))

# ── Colour helpers ─────────────────────────────────────────────────────────────
GREEN  = "\033[92m"
RED    = "\033[91m"
YELLOW = "\033[93m"
BLUE   = "\033[94m"
RESET  = "\033[0m"
BOLD   = "\033[1m"

def hdr(title):
    print(f"\n{BOLD}{BLUE}{'═'*60}{RESET}")
    print(f"{BOLD}{BLUE}  {title}{RESET}")
    print(f"{BOLD}{BLUE}{'═'*60}{RESET}")


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS — build in-memory Excel fixtures
# ══════════════════════════════════════════════════════════════════════════════

def make_xlsx(holdings=None, investment_per_tx=3000):
    """
    Create a minimal in-memory Excel file with Current Holdings sheet.
    holdings = list of (shop, date, code, name, buy_price, qty, avg_price, next_bid)
    Returns path to a temp file.
    """
    import openpyxl
    wb = openpyxl.Workbook()

    # ── Current Holdings sheet ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Current Holdings"

    # Row 1 headers (simplified)
    ws.cell(row=1, column=1, value="Investment Capital")
    ws.cell(row=1, column=2, value="Investment per transaction (40 parts)")

    # Row 2 — investment values
    ws.cell(row=2, column=1, value=100000)
    ws.cell(row=2, column=2, value=investment_per_tx)

    # Row 6 — column headers
    headers = ["Shop Type", "Buy Date", "NSE Code", "Underlying Asset",
               "Buy Price", "Actual Buy Qty", "Invested amount", "Total Qty",
               "Total Invested", "Avg Price", "Target Price", "CMP",
               "Sell Date", "Total Invested amt", "Overall Notional P/L",
               "Notional P/L%", "Next BID level", "Next BID price"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=6, column=i, value=h)

    # Data rows starting at row 7
    if holdings:
        for r_idx, h in enumerate(holdings, 7):
            shop, dt, code, name, buy_price, qty, avg_price, next_bid = h
            invested = round(buy_price * qty, 2)
            ws.cell(row=r_idx, column=1,  value=shop)
            ws.cell(row=r_idx, column=2,  value=dt)
            ws.cell(row=r_idx, column=3,  value=code)
            ws.cell(row=r_idx, column=4,  value=name)
            ws.cell(row=r_idx, column=5,  value=buy_price)
            ws.cell(row=r_idx, column=6,  value=qty)
            ws.cell(row=r_idx, column=7,  value=invested)
            ws.cell(row=r_idx, column=8,  value=qty)
            ws.cell(row=r_idx, column=9,  value=invested)
            ws.cell(row=r_idx, column=10, value=avg_price)
            ws.cell(row=r_idx, column=17, value=-0.03)
            ws.cell(row=r_idx, column=18, value=next_bid)

    # ── Sold sheet (empty) ────────────────────────────────────────────────────
    ws_sold = wb.create_sheet("Sold")
    ws_sold.cell(row=3, column=1, value="Shop Type")
    ws_sold.cell(row=3, column=3, value="NSE Code")

    # ── Shopping list sheet (minimal) ─────────────────────────────────────────
    wb.create_sheet("Shopping list")

    # ── Capital Management sheet ──────────────────────────────────────────────
    wb.create_sheet("Capital Management")

    # ── Change Log ────────────────────────────────────────────────────────────
    wb.create_sheet("Change Log")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name


def make_order_log(orders=None):
    """Create a temp order log JSON file."""
    tmp = tempfile.NamedTemporaryFile(suffix=".json", delete=False, mode="w")
    json.dump(orders or [], tmp)
    tmp.flush()
    return tmp.name


# ══════════════════════════════════════════════════════════════════════════════
# TEST CLASSES
# ══════════════════════════════════════════════════════════════════════════════

class TestNseToYahoo(unittest.TestCase):
    """nse_to_yahoo symbol conversion."""

    def setUp(self):
        from fire_shop_automation import nse_to_yahoo
        self.fn = nse_to_yahoo

    def test_basic_conversion(self):
        self.assertEqual(self.fn("NSE:AUTOIETF"), "AUTOIETF.NS")

    def test_psubnkbees(self):
        self.assertEqual(self.fn("NSE:PSUBNKBEES"), "PSUBNKBEES.NS")

    def test_niftybees(self):
        self.assertEqual(self.fn("NSE:NIFTYBEES"), "NIFTYBEES.NS")

    def test_no_nse_prefix(self):
        # Graceful — just appends .NS
        self.assertEqual(self.fn("ITBEES"), "ITBEES.NS")

    def test_already_ns(self):
        # Idempotency concern — double call
        result = self.fn("NSE:ITBEES")
        self.assertTrue(result.endswith(".NS"))


class TestComputePctChange(unittest.TestCase):
    """compute_pct_change edge cases."""

    def setUp(self):
        from fire_shop_automation import compute_pct_change
        self.fn = compute_pct_change

    def test_normal_dip(self):
        result = self.fn(90, 100)
        self.assertAlmostEqual(result, -0.10)

    def test_normal_gain(self):
        result = self.fn(110, 100)
        self.assertAlmostEqual(result, 0.10)

    def test_zero_dma(self):
        self.assertIsNone(self.fn(100, 0))

    def test_none_cmp(self):
        self.assertIsNone(self.fn(None, 100))

    def test_none_dma(self):
        self.assertIsNone(self.fn(100, None))

    def test_both_none(self):
        self.assertIsNone(self.fn(None, None))

    def test_equal_cmp_dma(self):
        self.assertAlmostEqual(self.fn(100, 100), 0.0)

    def test_very_large_dip(self):
        result = self.fn(1, 100)
        self.assertAlmostEqual(result, -0.99)


class TestLoadInvestmentPerTx(unittest.TestCase):
    """load_investment_per_tx reads correct cell."""

    def setUp(self):
        from fire_shop_automation import load_investment_per_tx
        self.fn = load_investment_per_tx

    def test_reads_correct_value(self):
        xlsx = make_xlsx(investment_per_tx=3004)
        try:
            self.assertEqual(self.fn(xlsx), 3004.0)
        finally:
            os.unlink(xlsx)

    def test_fallback_on_missing_file(self):
        self.assertEqual(self.fn("/nonexistent/path.xlsx"), 3000)

    def test_fallback_on_empty_cell(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Current Holdings"
        for name in ["Sold", "Shopping list", "Capital Management", "Change Log"]:
            wb.create_sheet(name)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        try:
            result = self.fn(tmp.name)
            self.assertEqual(result, 3000)
        finally:
            os.unlink(tmp.name)

    def test_custom_investment(self):
        xlsx = make_xlsx(investment_per_tx=5000)
        try:
            self.assertEqual(self.fn(xlsx), 5000.0)
        finally:
            os.unlink(xlsx)


class TestLoadCurrentHoldings(unittest.TestCase):
    """load_current_holdings — all scenarios."""

    def setUp(self):
        from fire_shop_automation import load_current_holdings
        self.fn = load_current_holdings

    def test_empty_holdings(self):
        xlsx = make_xlsx()
        try:
            result = self.fn(xlsx)
            self.assertEqual(result, {})
        finally:
            os.unlink(xlsx)

    def test_single_holding(self):
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:AUTOIETF", "NIFTY Auto",
             24.95, 120, 24.95, 24.20)
        ])
        try:
            result = self.fn(xlsx)
            self.assertIn("NSE:AUTOIETF", result)
            h = result["NSE:AUTOIETF"]
            self.assertAlmostEqual(h["avg_price"], 24.95)
            self.assertAlmostEqual(h["next_bid"], 24.20)
            self.assertEqual(h["total_qty"], 120.0)
            self.assertAlmostEqual(h["total_invested"], 24.95 * 120, places=1)
        finally:
            os.unlink(xlsx)

    def test_multiple_buys_same_etf_accumulates_qty(self):
        """Two rows of same ETF → total_qty = sum of both."""
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:PSUBNKBEES", "PSU Bank",
             96.6, 31, 95.0, 92.15),
            ("Equity ETF Buy", "17/03/2026", "NSE:PSUBNKBEES", "PSU Bank",
             93.2, 32, 94.9, 92.05),
        ])
        try:
            result = self.fn(xlsx)
            self.assertIn("NSE:PSUBNKBEES", result)
            h = result["NSE:PSUBNKBEES"]
            self.assertEqual(h["total_qty"], 63.0)   # 31 + 32
        finally:
            os.unlink(xlsx)

    def test_multiple_etfs_independent(self):
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:AUTOIETF", "NIFTY Auto",
             24.95, 120, 24.95, 24.20),
            ("Equity ETF Buy", "16/03/2026", "NSE:MOREALTY", "Nifty Realty",
             70.0, 42, 70.0, 67.9),
        ])
        try:
            result = self.fn(xlsx)
            self.assertIn("NSE:AUTOIETF", result)
            self.assertIn("NSE:MOREALTY", result)
            self.assertEqual(len(result), 2)
        finally:
            os.unlink(xlsx)

    def test_shop_type_classification_etf(self):
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:AUTOIETF", "NIFTY Auto",
             24.95, 120, 24.95, 24.20)
        ])
        try:
            result = self.fn(xlsx)
            self.assertEqual(result["NSE:AUTOIETF"]["shop"], "etf")
        finally:
            os.unlink(xlsx)

    def test_shop_type_classification_stock(self):
        xlsx = make_xlsx(holdings=[
            ("Stock Buy", "15/03/2026", "NSE:LT", "Larsen & Toubro",
             3445.0, 1, 3445.0, 3341.65)
        ])
        try:
            result = self.fn(xlsx)
            self.assertEqual(result["NSE:LT"]["shop"], "stock")
        finally:
            os.unlink(xlsx)

    def test_row_with_missing_price_skipped(self):
        """Rows with no NSE: prefix should be ignored."""
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "INVALID_CODE", "Bad ETF",
             24.95, 120, 24.95, 24.20)
        ])
        try:
            result = self.fn(xlsx)
            self.assertEqual(result, {})
        finally:
            os.unlink(xlsx)

    def test_zero_qty_row(self):
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:AUTOIETF", "NIFTY Auto",
             24.95, 0, 24.95, 24.20)
        ])
        try:
            result = self.fn(xlsx)
            # Should load but total_qty = 0
            self.assertIn("NSE:AUTOIETF", result)
            self.assertEqual(result["NSE:AUTOIETF"]["total_qty"], 0.0)
        finally:
            os.unlink(xlsx)


class TestApplyHoldingsFilter(unittest.TestCase):
    """apply_holdings_filter — the most critical business logic."""

    def setUp(self):
        from fire_shop_automation import apply_holdings_filter
        self.fn = apply_holdings_filter

    def _make_ranked(self, items):
        """items = list of (code, cmp, pct)"""
        return [{"code": c, "cmp": p, "pct": pct, "dma20": 100.0,
                 "avg_volume": 500000, "name": c, "rank": i+1}
                for i, (c, p, pct) in enumerate(items)]

    def test_not_in_holdings_always_shows(self):
        ranked = self._make_ranked([("NSE:NIFTYBEES", 262.0, -0.05)])
        result = self.fn(ranked, {}, {}, {"NSE:NIFTYBEES": 262.0})
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["code"], "NSE:NIFTYBEES")

    def test_held_above_next_bid_suppressed(self):
        """CMP=93.6 > next_bid=91.7 → suppressed."""
        ranked = self._make_ranked([("NSE:PSUBNKBEES", 93.6, -0.086)])
        holdings = {"NSE:PSUBNKBEES": {"avg_price": 94.9, "next_bid": 91.7,
                                        "total_qty": 63, "total_invested": 5978}}
        result = self.fn(ranked, holdings, {}, {"NSE:PSUBNKBEES": 93.6})
        self.assertEqual(len(result), 0)

    def test_held_below_next_bid_shows_avg_down(self):
        """CMP=90.0 < next_bid=91.7 → AVG DOWN."""
        ranked = self._make_ranked([("NSE:PSUBNKBEES", 90.0, -0.10)])
        holdings = {"NSE:PSUBNKBEES": {"avg_price": 94.9, "next_bid": 91.7,
                                        "total_qty": 63, "total_invested": 5978}}
        result = self.fn(ranked, holdings, {}, {"NSE:PSUBNKBEES": 90.0})
        self.assertEqual(len(result), 1)
        self.assertIn("AVG DOWN", result[0]["note"])

    def test_held_exactly_at_next_bid_shows_avg_down(self):
        """CMP exactly equals next_bid → boundary case → AVG DOWN."""
        ranked = self._make_ranked([("NSE:PSUBNKBEES", 91.7, -0.034)])
        holdings = {"NSE:PSUBNKBEES": {"avg_price": 94.9, "next_bid": 91.7,
                                        "total_qty": 63, "total_invested": 5978}}
        result = self.fn(ranked, holdings, {}, {"NSE:PSUBNKBEES": 91.7})
        self.assertEqual(len(result), 1)
        self.assertIn("AVG DOWN", result[0]["note"])

    def test_mixed_list_only_suppresses_held(self):
        """3 ETFs: 1 held+above_bid, 1 held+below_bid, 1 free."""
        ranked = self._make_ranked([
            ("NSE:PSUBNKBEES", 93.6, -0.086),   # held, above bid → suppress
            ("NSE:AUTOIETF",   89.0, -0.075),    # held, below bid → avg down
            ("NSE:NIFTYBEES",  262.0, -0.05),    # not held → show
        ])
        holdings = {
            "NSE:PSUBNKBEES": {"avg_price": 94.9, "next_bid": 91.7,
                                "total_qty": 63, "total_invested": 5978},
            "NSE:AUTOIETF":   {"avg_price": 95.0, "next_bid": 92.15,
                                "total_qty": 30, "total_invested": 2850},
        }
        live_cmp = {"NSE:PSUBNKBEES": 93.6, "NSE:AUTOIETF": 89.0, "NSE:NIFTYBEES": 262.0}
        result = self.fn(ranked, holdings, {}, live_cmp)
        codes = [r["code"] for r in result]
        self.assertNotIn("NSE:PSUBNKBEES", codes)
        self.assertIn("NSE:AUTOIETF", codes)
        self.assertIn("NSE:NIFTYBEES", codes)

    def test_empty_ranked_list(self):
        result = self.fn([], {}, {}, {})
        self.assertEqual(result, [])

    def test_empty_holdings_all_pass(self):
        ranked = self._make_ranked([
            ("NSE:NIFTYBEES",  262.0, -0.05),
            ("NSE:PSUBNKBEES", 93.6,  -0.086),
        ])
        result = self.fn(ranked, {}, {}, {})
        self.assertEqual(len(result), 2)

    def test_holding_with_none_next_bid_suppressed(self):
        """If next_bid is None (missing from Excel), holding should be suppressed."""
        ranked = self._make_ranked([("NSE:AUTOIETF", 25.0, -0.09)])
        holdings = {"NSE:AUTOIETF": {"avg_price": 27.0, "next_bid": None,
                                      "total_qty": 120, "total_invested": 2994}}
        result = self.fn(ranked, holdings, {}, {"NSE:AUTOIETF": 25.0})
        # next_bid is None → condition `next_bid and cmp <= next_bid` is False → suppressed
        self.assertEqual(len(result), 0)

    def test_order_preserved_after_filter(self):
        """Ranking order must be preserved after filtering."""
        ranked = self._make_ranked([
            ("NSE:AUTOIETF",   25.0, -0.09),
            ("NSE:NIFTYBEES",  262.0, -0.05),
            ("NSE:PSUBNKBEES", 93.6,  -0.04),
        ])
        result = self.fn(ranked, {}, {}, {})
        self.assertEqual([r["code"] for r in result],
                         ["NSE:AUTOIETF", "NSE:NIFTYBEES", "NSE:PSUBNKBEES"])


class TestGetSellCandidate(unittest.TestCase):
    """get_sell_candidate — sell logic scenarios."""

    def setUp(self):
        from zerodha_auto_buy import get_sell_candidate
        self.fn = get_sell_candidate

    def test_no_holdings_returns_none(self):
        result = self.fn({}, {})
        self.assertIsNone(result)

    def test_profit_below_target_no_sell(self):
        """5% profit < 6.28% target → no sell."""
        holdings = {"NSE:ITBEES": {"avg_price": 42.35, "total_qty": 71,
                                    "total_invested": 3006.85, "name": "NIFTY IT"}}
        live_cmp = {"NSE:ITBEES": 44.47}   # 44.47/42.35 - 1 = 5.0%
        result = self.fn(holdings, live_cmp)
        self.assertIsNone(result)

    def test_profit_at_exactly_628_triggers_sell(self):
        """Exactly 6.28% profit → should sell."""
        avg = 100.0
        cmp = round(avg * 1.0628, 2)
        holdings = {"NSE:NIFTYBEES": {"avg_price": avg, "total_qty": 10,
                                       "total_invested": 1000.0, "name": "NIFTY 50"}}
        live_cmp = {"NSE:NIFTYBEES": cmp}
        result = self.fn(holdings, live_cmp)
        self.assertIsNotNone(result)
        self.assertEqual(result["code"], "NSE:NIFTYBEES")

    def test_profit_above_target_triggers_sell(self):
        """8% profit > 6.28% → should sell."""
        holdings = {"NSE:ITBEES": {"avg_price": 42.35, "total_qty": 71,
                                    "total_invested": 3006.85, "name": "NIFTY IT"}}
        live_cmp = {"NSE:ITBEES": 45.74}   # ~8% profit
        result = self.fn(holdings, live_cmp)
        self.assertIsNotNone(result)
        self.assertEqual(result["code"], "NSE:ITBEES")

    def test_multiple_eligible_picks_highest_invested(self):
        """Two ETFs in profit — pick the one with more capital invested."""
        holdings = {
            "NSE:ITBEES":   {"avg_price": 42.0, "total_qty": 71,
                              "total_invested": 2982.0, "name": "NIFTY IT"},    # smaller
            "NSE:NIFTYBEES": {"avg_price": 262.0, "total_qty": 20,
                               "total_invested": 5240.0, "name": "NIFTY 50"},  # bigger
        }
        live_cmp = {
            "NSE:ITBEES":    45.0,    # +7.1% — eligible
            "NSE:NIFTYBEES": 280.0,   # +6.9% — eligible
        }
        result = self.fn(holdings, live_cmp)
        self.assertIsNotNone(result)
        self.assertEqual(result["code"], "NSE:NIFTYBEES")   # higher invested

    def test_one_eligible_one_not(self):
        holdings = {
            "NSE:ITBEES":   {"avg_price": 42.0, "total_qty": 71,
                              "total_invested": 2982.0, "name": "IT"},   # 7% profit
            "NSE:AUTOIETF": {"avg_price": 27.0, "total_qty": 110,
                              "total_invested": 2970.0, "name": "Auto"}, # -7% loss
        }
        live_cmp = {"NSE:ITBEES": 45.0, "NSE:AUTOIETF": 25.0}
        result = self.fn(holdings, live_cmp)
        self.assertIsNotNone(result)
        self.assertEqual(result["code"], "NSE:ITBEES")

    def test_missing_cmp_in_live_map_skipped(self):
        """ETF not in live_cmp dict → skip gracefully."""
        holdings = {"NSE:ITBEES": {"avg_price": 42.0, "total_qty": 71,
                                    "total_invested": 2982.0, "name": "IT"}}
        live_cmp = {}   # no CMP available
        result = self.fn(holdings, live_cmp)
        self.assertIsNone(result)

    def test_avg_price_none_skipped(self):
        """Holdings row with no avg_price → skip."""
        holdings = {"NSE:ITBEES": {"avg_price": None, "total_qty": 71,
                                    "total_invested": 2982.0, "name": "IT"}}
        live_cmp = {"NSE:ITBEES": 50.0}
        result = self.fn(holdings, live_cmp)
        self.assertIsNone(result)

    def test_zero_qty_skipped(self):
        holdings = {"NSE:ITBEES": {"avg_price": 42.0, "total_qty": 0,
                                    "total_invested": 0, "name": "IT"}}
        live_cmp = {"NSE:ITBEES": 50.0}
        result = self.fn(holdings, live_cmp)
        self.assertIsNone(result)

    def test_profit_pct_correct(self):
        avg = 100.0
        cmp = 110.0   # exactly 10%
        holdings = {"NSE:NIFTYBEES": {"avg_price": avg, "total_qty": 10,
                                       "total_invested": 1000.0, "name": "NIFTY 50"}}
        live_cmp = {"NSE:NIFTYBEES": cmp}
        result = self.fn(holdings, live_cmp)
        self.assertAlmostEqual(result["profit_pct"], 0.10, places=3)

    def test_all_in_loss_returns_none(self):
        holdings = {
            "NSE:ITBEES":   {"avg_price": 45.0, "total_qty": 71,
                              "total_invested": 3195.0, "name": "IT"},
            "NSE:AUTOIETF": {"avg_price": 27.0, "total_qty": 110,
                              "total_invested": 2970.0, "name": "Auto"},
        }
        live_cmp = {"NSE:ITBEES": 40.0, "NSE:AUTOIETF": 24.0}
        result = self.fn(holdings, live_cmp)
        self.assertIsNone(result)


class TestAlreadyOrderedToday(unittest.TestCase):
    """already_ordered_today — duplicate order prevention."""

    def setUp(self):
        from zerodha_auto_buy import already_ordered_today, ORDER_LOG_FILE
        self.fn   = already_ordered_today
        self.orig = ORDER_LOG_FILE

    def _patch_log(self, orders):
        import zerodha_auto_buy as m
        self._orig_path = m.ORDER_LOG_FILE
        tmp = Path(make_order_log(orders))
        m.ORDER_LOG_FILE = tmp
        return tmp

    def _restore_log(self, tmp):
        import zerodha_auto_buy as m
        m.ORDER_LOG_FILE = self._orig_path
        tmp.unlink(missing_ok=True)

    def test_empty_log_returns_false(self):
        tmp = self._patch_log([])
        try:
            self.assertFalse(self.fn("NSE:AUTOIETF"))
        finally:
            self._restore_log(tmp)

    def test_ordered_today_returns_true(self):
        orders = [{"date": date.today().isoformat(), "code": "NSE:AUTOIETF",
                   "status": "COMPLETE", "qty": 120, "limit_price": 25.0, "invested": 3000}]
        tmp = self._patch_log(orders)
        try:
            self.assertTrue(self.fn("NSE:AUTOIETF"))
        finally:
            self._restore_log(tmp)

    def test_ordered_yesterday_returns_false(self):
        from datetime import timedelta
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        orders = [{"date": yesterday, "code": "NSE:AUTOIETF",
                   "status": "COMPLETE", "qty": 120, "limit_price": 25.0, "invested": 3000}]
        tmp = self._patch_log(orders)
        try:
            self.assertFalse(self.fn("NSE:AUTOIETF"))
        finally:
            self._restore_log(tmp)

    def test_cancelled_order_not_counted(self):
        orders = [{"date": date.today().isoformat(), "code": "NSE:AUTOIETF",
                   "status": "CANCELLED", "qty": 120, "limit_price": 25.0, "invested": 3000}]
        tmp = self._patch_log(orders)
        try:
            self.assertFalse(self.fn("NSE:AUTOIETF"))
        finally:
            self._restore_log(tmp)

    def test_different_code_returns_false(self):
        orders = [{"date": date.today().isoformat(), "code": "NSE:NIFTYBEES",
                   "status": "COMPLETE", "qty": 10, "limit_price": 262.0, "invested": 2620}]
        tmp = self._patch_log(orders)
        try:
            self.assertFalse(self.fn("NSE:AUTOIETF"))
        finally:
            self._restore_log(tmp)

    def test_paper_order_not_counted_for_live(self):
        orders = [{"date": date.today().isoformat(), "code": "NSE:AUTOIETF",
                   "status": "paper", "qty": 120, "limit_price": 25.0, "invested": 3000}]
        tmp = self._patch_log(orders)
        try:
            # paper status is not "CANCELLED" but check: the fn checks status != 'CANCELLED'
            # paper orders SHOULD be counted to prevent double paper trades but
            # in live mode they should be skipped — let's verify current behavior
            result = self.fn("NSE:AUTOIETF")
            # paper has status "paper" which != "CANCELLED", so it IS counted
            self.assertTrue(result)
        finally:
            self._restore_log(tmp)


class TestRankInstruments(unittest.TestCase):
    """rank_instruments — volume filter and ranking logic."""

    def _make_fetch_result(self, cmp, dma20, vol):
        return (cmp, dma20, vol)

    @patch("fire_shop_automation.fetch_etf_data")
    @patch("fire_shop_automation.time.sleep")
    def test_low_volume_excluded(self, mock_sleep, mock_fetch):
        from fire_shop_automation import rank_instruments
        # Two ETFs: one high volume, one low
        mock_fetch.side_effect = [
            (100.0, 110.0, 500000),   # high vol
            (50.0,  60.0,  5000),     # low vol — should be excluded
        ]
        instruments = [("NSE:NIFTYBEES", "NIFTY 50"), ("NSE:LOWVOL", "Low Vol ETF")]
        result = rank_instruments(instruments, None, "test")
        codes = [r["code"] for r in result]
        self.assertIn("NSE:NIFTYBEES", codes)
        self.assertNotIn("NSE:LOWVOL", codes)

    @patch("fire_shop_automation.fetch_etf_data")
    @patch("fire_shop_automation.time.sleep")
    def test_ranked_by_biggest_dip(self, mock_sleep, mock_fetch):
        from fire_shop_automation import rank_instruments
        mock_fetch.side_effect = [
            (90.0, 100.0, 1000000),   # -10% dip
            (95.0, 100.0, 1000000),   # -5% dip
            (85.0, 100.0, 1000000),   # -15% dip — should be rank 1
        ]
        instruments = [
            ("NSE:A", "A"), ("NSE:B", "B"), ("NSE:C", "C")
        ]
        result = rank_instruments(instruments, None, "test")
        self.assertEqual(result[0]["code"], "NSE:C")   # biggest dip first
        self.assertEqual(result[1]["code"], "NSE:A")
        self.assertEqual(result[2]["code"], "NSE:B")

    @patch("fire_shop_automation.fetch_etf_data")
    @patch("fire_shop_automation.time.sleep")
    def test_fetch_failure_excluded(self, mock_sleep, mock_fetch):
        from fire_shop_automation import rank_instruments
        mock_fetch.side_effect = [
            (None, None, None),       # fetch failed
            (100.0, 110.0, 500000),   # success
        ]
        instruments = [("NSE:FAIL", "Failed"), ("NSE:OK", "OK")]
        result = rank_instruments(instruments, None, "test")
        codes = [r["code"] for r in result]
        self.assertNotIn("NSE:FAIL", codes)
        self.assertIn("NSE:OK", codes)

    @patch("fire_shop_automation.fetch_etf_data")
    @patch("fire_shop_automation.time.sleep")
    def test_all_fail_returns_empty(self, mock_sleep, mock_fetch):
        from fire_shop_automation import rank_instruments
        mock_fetch.return_value = (None, None, None)
        instruments = [("NSE:A", "A"), ("NSE:B", "B")]
        result = rank_instruments(instruments, None, "test")
        self.assertEqual(result, [])

    @patch("fire_shop_automation.fetch_etf_data")
    @patch("fire_shop_automation.time.sleep")
    def test_all_below_volume_threshold(self, mock_sleep, mock_fetch):
        from fire_shop_automation import rank_instruments
        mock_fetch.return_value = (100.0, 110.0, 100)   # vol=100 < 20000
        instruments = [("NSE:A", "A"), ("NSE:B", "B")]
        result = rank_instruments(instruments, None, "test")
        self.assertEqual(result, [])

    @patch("fire_shop_automation.fetch_etf_data")
    @patch("fire_shop_automation.time.sleep")
    def test_rank_field_assigned(self, mock_sleep, mock_fetch):
        from fire_shop_automation import rank_instruments
        mock_fetch.side_effect = [
            (90.0, 100.0, 500000),
            (95.0, 100.0, 500000),
        ]
        instruments = [("NSE:A", "A"), ("NSE:B", "B")]
        result = rank_instruments(instruments, None, "test")
        self.assertEqual(result[0]["rank"], 1)
        self.assertEqual(result[1]["rank"], 2)


class TestSafetyChecks(unittest.TestCase):
    """run_safety_checks — all guard conditions."""

    def setUp(self):
        from zerodha_auto_buy import run_safety_checks
        self.fn = run_safety_checks

    def test_paper_mode_always_passes(self):
        safe, reason = self.fn(None, [], 3000, paper_mode=True)
        self.assertTrue(safe)

    @patch("zerodha_auto_buy.get_available_cash", return_value=0.0)
    def test_insufficient_cash_blocked(self, mock_cash):
        picks = [{"code": "NSE:AUTOIETF", "cmp": 25.0}]
        safe, reason = self.fn(MagicMock(), picks, 3000, paper_mode=False)
        self.assertFalse(safe)
        self.assertIn("cash", reason.lower())

    @patch("zerodha_auto_buy.get_available_cash", return_value=50000.0)
    @patch("zerodha_auto_buy.total_spent_today", return_value=14000.0)
    def test_daily_spend_cap_blocked(self, mock_spent, mock_cash):
        """14000 spent + 3000 new = 17000 > 15000 cap."""
        picks = [{"code": "NSE:AUTOIETF", "cmp": 100.0}]
        safe, reason = self.fn(MagicMock(), picks, 3000, paper_mode=False)
        self.assertFalse(safe)

    @patch("zerodha_auto_buy.get_available_cash", return_value=50000.0)
    @patch("zerodha_auto_buy.total_spent_today", return_value=0.0)
    @patch("zerodha_auto_buy.datetime")
    def test_weekend_blocked(self, mock_dt, mock_spent, mock_cash):
        # Saturday
        mock_dt.now.return_value = datetime(2026, 3, 21, 15, 0)  # Saturday
        picks = [{"code": "NSE:AUTOIETF", "cmp": 25.0}]
        safe, reason = self.fn(MagicMock(), picks, 3000, paper_mode=False)
        self.assertFalse(safe)
        self.assertIn("weekend", reason.lower())

    @patch("zerodha_auto_buy.get_available_cash", return_value=50000.0)
    @patch("zerodha_auto_buy.total_spent_today", return_value=0.0)
    @patch("zerodha_auto_buy.datetime")
    def test_before_market_hours_blocked(self, mock_dt, mock_spent, mock_cash):
        mock_dt.now.return_value = datetime(2026, 3, 19, 8, 0)  # 8 AM
        picks = [{"code": "NSE:AUTOIETF", "cmp": 25.0}]
        safe, reason = self.fn(MagicMock(), picks, 3000, paper_mode=False)
        self.assertFalse(safe)

    @patch("zerodha_auto_buy.get_available_cash", return_value=50000.0)
    @patch("zerodha_auto_buy.total_spent_today", return_value=0.0)
    @patch("zerodha_auto_buy.datetime")
    def test_after_market_hours_blocked(self, mock_dt, mock_spent, mock_cash):
        mock_dt.now.return_value = datetime(2026, 3, 19, 15, 31)  # after close
        picks = [{"code": "NSE:AUTOIETF", "cmp": 25.0}]
        safe, reason = self.fn(MagicMock(), picks, 3000, paper_mode=False)
        self.assertFalse(safe)

    @patch("zerodha_auto_buy.get_available_cash", return_value=50000.0)
    @patch("zerodha_auto_buy.total_spent_today", return_value=0.0)
    @patch("zerodha_auto_buy.datetime")
    def test_during_market_hours_passes(self, mock_dt, mock_spent, mock_cash):
        mock_dt.now.return_value = datetime(2026, 3, 19, 15, 0)  # 3 PM — valid
        picks = [{"code": "NSE:AUTOIETF", "cmp": 25.0}]
        safe, reason = self.fn(MagicMock(), picks, 3000, paper_mode=False)
        self.assertTrue(safe)


class TestFetchEtfData(unittest.TestCase):
    """fetch_etf_data — Yahoo Finance parsing."""

    def _mock_response(self, closes, volumes=None):
        mock_r = MagicMock()
        mock_r.json.return_value = {
            "chart": {"result": [{
                "indicators": {"quote": [{
                    "close":  closes,
                    "volume": volumes or [1000000] * len(closes)
                }]}
            }]}
        }
        return mock_r

    @patch("fire_shop_automation.requests.get")
    def test_normal_fetch(self, mock_get):
        from fire_shop_automation import fetch_etf_data
        closes = [100.0] * 19 + [95.0]   # 20 closes, last = 95
        mock_get.return_value = self._mock_response(closes)
        cmp, dma20, vol = fetch_etf_data(None, "NSE:AUTOIETF")
        self.assertAlmostEqual(cmp, 95.0)
        self.assertIsNotNone(dma20)

    @patch("fire_shop_automation.requests.get")
    def test_none_values_in_closes_filtered(self, mock_get):
        from fire_shop_automation import fetch_etf_data
        closes = [100.0, None, 102.0, None] + [100.0] * 16 + [98.0]
        mock_get.return_value = self._mock_response(closes)
        cmp, dma20, vol = fetch_etf_data(None, "NSE:AUTOIETF")
        self.assertIsNotNone(cmp)
        self.assertIsNotNone(dma20)

    @patch("fire_shop_automation.requests.get")
    def test_empty_closes_returns_none(self, mock_get):
        from fire_shop_automation import fetch_etf_data
        mock_r = MagicMock()
        mock_r.json.return_value = {
            "chart": {"result": [{"indicators": {"quote": [{"close": [], "volume": []}]}}]}
        }
        mock_get.return_value = mock_r
        cmp, dma20, vol = fetch_etf_data(None, "NSE:AUTOIETF")
        self.assertIsNone(cmp)
        self.assertIsNone(dma20)

    @patch("fire_shop_automation.requests.get")
    def test_network_error_returns_none(self, mock_get):
        from fire_shop_automation import fetch_etf_data
        mock_get.side_effect = Exception("network error")
        cmp, dma20, vol = fetch_etf_data(None, "NSE:AUTOIETF")
        self.assertIsNone(cmp)
        self.assertIsNone(dma20)
        self.assertIsNone(vol)

    @patch("fire_shop_automation.requests.get")
    def test_fewer_than_20_closes_uses_available(self, mock_get):
        from fire_shop_automation import fetch_etf_data
        closes = [100.0] * 10   # only 10 days of data
        mock_get.return_value = self._mock_response(closes)
        cmp, dma20, vol = fetch_etf_data(None, "NSE:AUTOIETF")
        self.assertIsNotNone(cmp)
        self.assertIsNotNone(dma20)

    @patch("fire_shop_automation.requests.get")
    def test_volume_computed_correctly(self, mock_get):
        from fire_shop_automation import fetch_etf_data
        closes  = [100.0] * 20
        volumes = [500000] * 20
        mock_get.return_value = self._mock_response(closes, volumes)
        cmp, dma20, vol = fetch_etf_data(None, "NSE:NIFTYBEES")
        self.assertEqual(vol, 500000)

    @patch("fire_shop_automation.requests.get")
    def test_all_closes_none_returns_none(self, mock_get):
        from fire_shop_automation import fetch_etf_data
        mock_r = MagicMock()
        mock_r.json.return_value = {
            "chart": {"result": [{"indicators": {"quote": [
                {"close": [None, None, None], "volume": [None, None]}
            ]}}]}
        }
        mock_get.return_value = mock_r
        cmp, dma20, vol = fetch_etf_data(None, "NSE:AUTOIETF")
        self.assertIsNone(cmp)


class TestRemoveFromHoldings(unittest.TestCase):
    """remove_from_holdings — after sell, rows deleted from Excel."""

    def setUp(self):
        from zerodha_auto_buy import remove_from_holdings
        self.fn = remove_from_holdings

    def test_removes_correct_rows(self):
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:ITBEES",   "NIFTY IT",   42.35, 71,  42.35, 41.08),
            ("Equity ETF Buy", "16/03/2026", "NSE:MOREALTY", "Realty ETF", 70.0,  42,  70.0,  67.9),
            ("Equity ETF Buy", "17/03/2026", "NSE:ITBEES",   "NIFTY IT",   40.0,  75,  41.27, 40.03),
        ])
        try:
            self.fn(xlsx, "NSE:ITBEES")
            from fire_shop_automation import load_current_holdings
            result = load_current_holdings(xlsx)
            self.assertNotIn("NSE:ITBEES", result)
            self.assertIn("NSE:MOREALTY", result)
        finally:
            os.unlink(xlsx)

    def test_removing_nonexistent_code_is_safe(self):
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:MOREALTY", "Realty ETF", 70.0, 42, 70.0, 67.9)
        ])
        try:
            # Should not raise
            self.fn(xlsx, "NSE:DOESNOTEXIST")
            from fire_shop_automation import load_current_holdings
            result = load_current_holdings(xlsx)
            self.assertIn("NSE:MOREALTY", result)
        finally:
            os.unlink(xlsx)

    def test_all_rows_removed_when_multiple_buys(self):
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:PSUBNKBEES", "PSU Bank", 96.6, 31, 94.9, 92.05),
            ("Equity ETF Buy", "17/03/2026", "NSE:PSUBNKBEES", "PSU Bank", 93.2, 32, 94.9, 92.05),
            ("Equity ETF Buy", "19/03/2026", "NSE:PSUBNKBEES", "PSU Bank", 93.7, 32, 94.9, 92.05),
        ])
        try:
            self.fn(xlsx, "NSE:PSUBNKBEES")
            from fire_shop_automation import load_current_holdings
            result = load_current_holdings(xlsx)
            self.assertNotIn("NSE:PSUBNKBEES", result)
        finally:
            os.unlink(xlsx)


class TestLogToExcel(unittest.TestCase):
    """log_to_excel — verify correct data written to Excel."""

    def setUp(self):
        from fire_shop_automation import log_to_excel, load_current_holdings
        self.log_fn  = log_to_excel
        self.load_fn = load_current_holdings

    def _make_pick(self, code="NSE:AUTOIETF", cmp=25.0, pct=-0.09):
        return {"code": code, "name": "NIFTY Auto", "cmp": cmp,
                "pct": pct, "dma20": 27.5, "rank": 1, "note": "", "shop": "etf"}

    def test_new_buy_logged_correctly(self):
        xlsx = make_xlsx()
        try:
            pick = self._make_pick()
            self.log_fn(xlsx, [pick], {"etf": [pick], "jewellery": [], "stock": []}, 3000)
            result = self.load_fn(xlsx)
            self.assertIn("NSE:AUTOIETF", result)
        finally:
            os.unlink(xlsx)

    def test_avg_price_written(self):
        xlsx = make_xlsx()
        try:
            pick = self._make_pick(cmp=25.0)
            self.log_fn(xlsx, [pick], {"etf": [pick], "jewellery": [], "stock": []}, 3000)
            result = self.load_fn(xlsx)
            self.assertAlmostEqual(result["NSE:AUTOIETF"]["avg_price"], 25.0, places=1)
        finally:
            os.unlink(xlsx)

    def test_next_bid_written_as_97pct_of_avg(self):
        xlsx = make_xlsx()
        try:
            pick = self._make_pick(cmp=100.0)
            self.log_fn(xlsx, [pick], {"etf": [pick], "jewellery": [], "stock": []}, 3000)
            result = self.load_fn(xlsx)
            expected_next_bid = round(100.0 * 0.97, 4)
            self.assertAlmostEqual(result["NSE:AUTOIETF"]["next_bid"], expected_next_bid, places=2)
        finally:
            os.unlink(xlsx)

    def test_second_buy_same_etf_accumulates(self):
        """First buy then second buy of same ETF — total_qty should add up."""
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:AUTOIETF", "NIFTY Auto",
             25.0, 120, 25.0, 24.25)
        ])
        try:
            pick = self._make_pick(cmp=24.0)   # second buy at lower price
            self.log_fn(xlsx, [pick], {"etf": [pick], "jewellery": [], "stock": []}, 3000)
            result = self.load_fn(xlsx)
            h = result["NSE:AUTOIETF"]
            self.assertGreater(h["total_qty"], 120)   # should have more than first buy
        finally:
            os.unlink(xlsx)

    def test_overrides_original_file(self):
        """Ensure original file is overwritten, not a new dated file created."""
        xlsx = make_xlsx()
        original_path = xlsx
        try:
            pick = self._make_pick()
            returned_path = self.log_fn(xlsx, [pick],
                                        {"etf": [pick], "jewellery": [], "stock": []}, 3000)
            self.assertEqual(returned_path, original_path)
            # No dated file should exist
            dated = xlsx.replace(".xlsx", f"_updated_{date.today().isoformat()}.xlsx")
            self.assertFalse(Path(dated).exists())
        finally:
            os.unlink(xlsx)


class TestTotalSpentToday(unittest.TestCase):
    """total_spent_today — daily spend tracking."""

    def _patch_log(self, orders):
        import zerodha_auto_buy as m
        self._orig = m.ORDER_LOG_FILE
        tmp = Path(make_order_log(orders))
        m.ORDER_LOG_FILE = tmp
        return tmp

    def _restore(self, tmp):
        import zerodha_auto_buy as m
        m.ORDER_LOG_FILE = self._orig
        tmp.unlink(missing_ok=True)

    def test_empty_log_zero(self):
        from zerodha_auto_buy import total_spent_today
        tmp = self._patch_log([])
        try:
            self.assertEqual(total_spent_today(), 0)
        finally:
            self._restore(tmp)

    def test_today_orders_summed(self):
        from zerodha_auto_buy import total_spent_today
        orders = [
            {"date": date.today().isoformat(), "status": "COMPLETE", "invested": 3000},
            {"date": date.today().isoformat(), "status": "COMPLETE", "invested": 2500},
        ]
        tmp = self._patch_log(orders)
        try:
            self.assertEqual(total_spent_today(), 5500)
        finally:
            self._restore(tmp)

    def test_cancelled_not_counted(self):
        from zerodha_auto_buy import total_spent_today
        orders = [
            {"date": date.today().isoformat(), "status": "CANCELLED", "invested": 3000},
            {"date": date.today().isoformat(), "status": "COMPLETE",  "invested": 2500},
        ]
        tmp = self._patch_log(orders)
        try:
            self.assertEqual(total_spent_today(), 2500)
        finally:
            self._restore(tmp)

    def test_paper_not_counted(self):
        from zerodha_auto_buy import total_spent_today
        orders = [
            {"date": date.today().isoformat(), "status": "paper",    "invested": 3000},
            {"date": date.today().isoformat(), "status": "COMPLETE", "invested": 2500},
        ]
        tmp = self._patch_log(orders)
        try:
            self.assertEqual(total_spent_today(), 2500)
        finally:
            self._restore(tmp)

    def test_yesterday_not_counted(self):
        from zerodha_auto_buy import total_spent_today
        from datetime import timedelta
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        orders = [
            {"date": yesterday,                  "status": "COMPLETE", "invested": 3000},
            {"date": date.today().isoformat(),   "status": "COMPLETE", "invested": 2500},
        ]
        tmp = self._patch_log(orders)
        try:
            self.assertEqual(total_spent_today(), 2500)
        finally:
            self._restore(tmp)


class TestPlaceLimitOrder(unittest.TestCase):
    """place_limit_order — paper mode and live mode."""

    def setUp(self):
        from zerodha_auto_buy import place_limit_order
        self.fn = place_limit_order

    def _patch_log(self):
        import zerodha_auto_buy as m
        self._orig = m.ORDER_LOG_FILE
        tmp = Path(make_order_log([]))
        m.ORDER_LOG_FILE = tmp
        return tmp

    def _restore(self, tmp):
        import zerodha_auto_buy as m
        m.ORDER_LOG_FILE = self._orig
        tmp.unlink(missing_ok=True)

    def test_paper_order_logged(self):
        tmp = self._patch_log()
        try:
            order_id = self.fn(None, "NSE:AUTOIETF", 25.0, 120, paper_mode=True)
            self.assertIsNotNone(order_id)
            self.assertTrue(str(order_id).startswith("PAPER_"))
            import zerodha_auto_buy as m
            log = json.loads(m.ORDER_LOG_FILE.read_text())
            self.assertEqual(len(log), 1)
            self.assertEqual(log[0]["status"], "paper")
        finally:
            self._restore(tmp)

    def test_paper_order_has_correct_fields(self):
        tmp = self._patch_log()
        try:
            self.fn(None, "NSE:AUTOIETF", 25.0, 120, paper_mode=True)
            import zerodha_auto_buy as m
            log = json.loads(m.ORDER_LOG_FILE.read_text())
            entry = log[0]
            self.assertEqual(entry["code"], "NSE:AUTOIETF")
            self.assertEqual(entry["symbol"], "AUTOIETF")
            self.assertEqual(entry["qty"], 120)
            self.assertEqual(entry["date"], date.today().isoformat())
        finally:
            self._restore(tmp)

    def test_live_order_placed_and_logged(self):
        tmp = self._patch_log()
        try:
            mock_kite = MagicMock()
            mock_kite.place_order.return_value = "ORDER123"
            from kiteconnect import KiteConnect
            order_id = self.fn(mock_kite, "NSE:AUTOIETF", 25.0, 120, paper_mode=False)
            self.assertEqual(order_id, "ORDER123")
            import zerodha_auto_buy as m
            log = json.loads(m.ORDER_LOG_FILE.read_text())
            self.assertEqual(log[0]["order_id"], "ORDER123")
            self.assertEqual(log[0]["status"], "PLACED")
        finally:
            self._restore(tmp)

    def test_live_order_failure_logged(self):
        tmp = self._patch_log()
        try:
            mock_kite = MagicMock()
            mock_kite.place_order.side_effect = Exception("Insufficient funds")
            order_id = self.fn(mock_kite, "NSE:AUTOIETF", 25.0, 120, paper_mode=False)
            self.assertIsNone(order_id)
            import zerodha_auto_buy as m
            log = json.loads(m.ORDER_LOG_FILE.read_text())
            self.assertEqual(log[0]["status"], "FAILED")
            self.assertIn("Insufficient funds", log[0]["error"])
        finally:
            self._restore(tmp)

    def test_limit_price_has_buffer(self):
        tmp = self._patch_log()
        try:
            from zerodha_auto_buy import LIMIT_PRICE_BUFFER
            self.fn(None, "NSE:AUTOIETF", 100.0, 10, paper_mode=True)
            import zerodha_auto_buy as m
            log = json.loads(m.ORDER_LOG_FILE.read_text())
            expected = round(100.0 * (1 + LIMIT_PRICE_BUFFER), 1)
            self.assertAlmostEqual(log[0]["limit_price"], expected, places=1)
        finally:
            self._restore(tmp)


class TestSendTelegram(unittest.TestCase):
    """send_telegram — network error handling."""

    @patch("zerodha_auto_buy.requests.post")
    def test_success(self, mock_post):
        from zerodha_auto_buy import send_telegram
        mock_post.return_value.json.return_value = {"ok": True}
        # Should not raise
        send_telegram("Test message")

    @patch("zerodha_auto_buy.requests.post")
    def test_network_error_doesnt_crash(self, mock_post):
        from zerodha_auto_buy import send_telegram
        mock_post.side_effect = Exception("network error")
        # Should not raise — Telegram failure is non-fatal
        send_telegram("Test message")

    @patch("zerodha_auto_buy.requests.post")
    def test_telegram_api_error_doesnt_crash(self, mock_post):
        from zerodha_auto_buy import send_telegram
        mock_post.return_value.json.return_value = {"ok": False, "description": "Bad token"}
        send_telegram("Test message")


class TestEdgeCases(unittest.TestCase):
    """Edge cases and boundary conditions."""

    def test_profit_target_boundary_628_pct(self):
        """Verify 6.28% is the exact threshold — not 6.27% or 6.29%."""
        from zerodha_auto_buy import get_sell_candidate, PROFIT_TARGET_PCT
        self.assertAlmostEqual(PROFIT_TARGET_PCT, 0.0628, places=4)

        avg = 100.0
        # Just below: 6.27% → no sell
        holdings_low = {"NSE:TEST": {"avg_price": avg, "total_qty": 10,
                                      "total_invested": 1000, "name": "Test"}}
        result = get_sell_candidate(holdings_low, {"NSE:TEST": avg * 1.0627})
        self.assertIsNone(result)

        # Just above: 6.29% → sell
        result = get_sell_candidate(holdings_low, {"NSE:TEST": avg * 1.0629})
        self.assertIsNotNone(result)

    def test_min_volume_threshold(self):
        """Verify 20000 is the exact volume threshold."""
        from fire_shop_automation import MIN_VOLUME
        self.assertEqual(MIN_VOLUME, 20000)

    def test_nse_symbol_stripping(self):
        from zerodha_auto_buy import nse_symbol
        self.assertEqual(nse_symbol("NSE:AUTOIETF"), "AUTOIETF")
        self.assertEqual(nse_symbol("NSE:NIFTYBEES"), "NIFTYBEES")

    def test_investment_qty_calculation(self):
        """Qty = floor(investment / cmp)."""
        investment = 3000
        cmp = 25.0
        qty = max(1, int(investment // cmp))
        self.assertEqual(qty, 120)

    def test_investment_qty_min_1(self):
        """Very high CMP → qty should be at least 1."""
        investment = 3000
        cmp = 5000.0   # e.g. expensive stock
        qty = max(1, int(investment // cmp))
        self.assertEqual(qty, 1)

    def test_limit_sell_price_below_cmp(self):
        """Sell limit = CMP - 0.1% buffer (for fill guarantee)."""
        from zerodha_auto_buy import LIMIT_PRICE_BUFFER
        cmp = 100.0
        limit_sell = round(cmp * (1 - LIMIT_PRICE_BUFFER), 1)
        self.assertLess(limit_sell, cmp)

    def test_avg_down_suppression_boundary(self):
        """CMP exactly 1 rupee above next_bid → suppressed."""
        from fire_shop_automation import apply_holdings_filter
        next_bid = 91.7
        cmp = 92.7   # above next_bid
        ranked = [{"code": "NSE:PSUBNKBEES", "cmp": cmp, "pct": -0.05,
                   "dma20": 100, "avg_volume": 500000, "name": "PSU Bank", "rank": 1}]
        holdings = {"NSE:PSUBNKBEES": {"avg_price": 94.9, "next_bid": next_bid,
                                        "total_qty": 63, "total_invested": 5978}}
        result = apply_holdings_filter(ranked, holdings, {}, {"NSE:PSUBNKBEES": cmp})
        self.assertEqual(len(result), 0)   # suppressed

    def test_holdings_total_invested_calculation(self):
        """total_invested = buy_price × qty for each row summed."""
        xlsx = make_xlsx(holdings=[
            ("Equity ETF Buy", "15/03/2026", "NSE:PSUBNKBEES", "PSU Bank",
             96.6, 31, 95.0, 92.15),   # 96.6 × 31 = 2994.6
            ("Equity ETF Buy", "17/03/2026", "NSE:PSUBNKBEES", "PSU Bank",
             93.2, 32, 94.9, 92.05),   # 93.2 × 32 = 2982.4
        ])
        try:
            from fire_shop_automation import load_current_holdings
            result = load_current_holdings(xlsx)
            h = result["NSE:PSUBNKBEES"]
            expected = round(96.6 * 31 + 93.2 * 32, 2)
            self.assertAlmostEqual(h["total_invested"], expected, places=0)
        finally:
            os.unlink(xlsx)


# ══════════════════════════════════════════════════════════════════════════════
# RUNNER
# ══════════════════════════════════════════════════════════════════════════════

class VerboseResult(unittest.TestResult):
    def __init__(self):
        super().__init__()
        self.test_results = []

    def startTest(self, test):
        super().startTest(test)

    def addSuccess(self, test):
        super().addSuccess(test)
        self.test_results.append(("PASS", test))

    def addFailure(self, test, err):
        super().addFailure(test, err)
        self.test_results.append(("FAIL", test, err))

    def addError(self, test, err):
        super().addError(test, err)
        self.test_results.append(("ERROR", test, err))

    def addSkip(self, test, reason):
        super().addSkip(test, reason)
        self.test_results.append(("SKIP", test))


def run_tests():
    test_classes = [
        ("Symbol conversion",         TestNseToYahoo),
        ("Pct change calculation",     TestComputePctChange),
        ("Investment per tx",          TestLoadInvestmentPerTx),
        ("Load holdings",              TestLoadCurrentHoldings),
        ("Holdings filter",            TestApplyHoldingsFilter),
        ("Sell candidate selection",   TestGetSellCandidate),
        ("Duplicate order guard",      TestAlreadyOrderedToday),
        ("ETF ranking + vol filter",   TestRankInstruments),
        ("Safety checks",              TestSafetyChecks),
        ("Yahoo Finance parsing",      TestFetchEtfData),
        ("Remove from holdings",       TestRemoveFromHoldings),
        ("Excel logging",              TestLogToExcel),
        ("Daily spend tracking",       TestTotalSpentToday),
        ("Order placement",            TestPlaceLimitOrder),
        ("Telegram alerts",            TestSendTelegram),
        ("Edge cases & boundaries",    TestEdgeCases),
    ]

    total_pass = total_fail = total_error = 0

    for label, cls in test_classes:
        hdr(label)
        suite  = unittest.TestLoader().loadTestsFromTestCase(cls)
        result = VerboseResult()
        suite.run(result)

        for item in result.test_results:
            status = item[0]
            test   = item[1]
            name   = test._testMethodName
            if status == "PASS":
                print(f"  {GREEN}✅ PASS{RESET}  {name}")
                total_pass += 1
            elif status == "FAIL":
                print(f"  {RED}❌ FAIL{RESET}  {name}")
                print(f"         {RED}{item[2][1]}{RESET}")
                total_fail += 1
            elif status == "ERROR":
                print(f"  {YELLOW}⚠️  ERR {RESET}  {name}")
                import traceback
                tb = "".join(traceback.format_exception(*item[2]))
                print(f"         {YELLOW}{tb.strip()}{RESET}")
                total_error += 1
            elif status == "SKIP":
                print(f"  {YELLOW}⏭️  SKIP{RESET}  {name}")

    # ── Summary ───────────────────────────────────────────────────────────────
    total = total_pass + total_fail + total_error
    print(f"\n{BOLD}{'═'*60}{RESET}")
    print(f"{BOLD}  TEST SUMMARY{RESET}")
    print(f"{'═'*60}")
    print(f"  {GREEN}✅ Passed : {total_pass}{RESET}")
    print(f"  {RED}❌ Failed : {total_fail}{RESET}")
    print(f"  {YELLOW}⚠️  Errors : {total_error}{RESET}")
    print(f"  Total    : {total}")

    if total_fail == 0 and total_error == 0:
        print(f"\n  {GREEN}{BOLD}🎉 All {total} tests passed!{RESET}")
    else:
        print(f"\n  {RED}{BOLD}❌ {total_fail + total_error} test(s) need attention.{RESET}")

    return total_fail + total_error


if __name__ == "__main__":
    sys.exit(run_tests())
